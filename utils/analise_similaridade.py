import openpyxl
from .similaridade import medir_similaridade

def analyze_similarities(planilha, limiar_similaridade):

    try:
        achados_fp_for_sim = []
        achados_semclin_fn_sim = []
        index_semclin = []
        index_prompt = []
        narrativa_anterior = ""

        for index in range(2, planilha.max_row + 1):
            n1 = ''
            n2 = ''
            if planilha[f'A{index}'].value:
                n1 = str(planilha[f'A{index}'].value)[:4]
            elif planilha[f'G{index}'].value:
                n2 = str(planilha[f'G{index}'].value)[:4]

            narrativa_atual = n1 if n1 else n2

            if narrativa_atual and (narrativa_atual != narrativa_anterior or index == planilha.max_row):
                if achados_semclin_fn_sim and achados_fp_for_sim:
                    matches = []
                    for i_p, t_prompt in enumerate(achados_fp_for_sim):
                        for i_s, t_semclin in enumerate(achados_semclin_fn_sim):
                            t_prompt_str = str(t_prompt)
                            t_semclin_str = str(t_semclin)

                            if not t_prompt_str or not t_semclin_str:
                                continue

                            resultado = medir_similaridade(t_prompt_str, t_semclin_str)

                            if resultado > limiar_similaridade:
                                print(f"\n{resultado:.3f} -> {t_prompt_str} + {t_semclin_str}")
                                matches.append((i_p, i_s))
                                break

                    for i_p, i_s in reversed(matches):

                        classificacao_atual = planilha[f'J{index_prompt[i_p]}'].value
                        if classificacao_atual in ['FN', 'FP']:
                            planilha[f'J{index_prompt[i_p]}'] = 'VPP'
                        planilha[f'J{index_semclin[i_s]}'] = ''

                        achados_fp_for_sim.pop(i_p)
                        achados_semclin_fn_sim.pop(i_s)
                        index_prompt.pop(i_p)
                        index_semclin.pop(i_s)

                achados_fp_for_sim = []
                achados_semclin_fn_sim = []
                index_semclin = []
                index_prompt = []

            if not (index == planilha.max_row and narrativa_atual == narrativa_anterior):
                avaliacao = planilha[f'J{index}'].value
                if avaliacao == 'FN':
                    termo = planilha[f'H{index}'].value
                    if termo is not None:
                        achados_semclin_fn_sim.append(str(termo))
                        index_semclin.append(index)
                elif avaliacao == 'FP':
                    termo = planilha[f'D{index}'].value
                    if termo is not None:
                        achados_fp_for_sim.append(str(termo))
                        index_prompt.append(index)

            narrativa_anterior = narrativa_atual
    except Exception as e:
        print(f"\nErro durante análise de similaridade: {e}")
        raise
